from __future__ import annotations

import os.path
import re

import utils.message as message
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from peakrdl.ipxact import IPXACTImporter
from systemrdl import RDLCompileError, RDLCompiler

from .excel.args import EXCEL_REG_FIELD, EXCEL_REG_HEAD


class ExcelParser:
    """
    Parse Excel Worksheets

    `check_format`: basic format check
    `check_name` : register name and abbreviation check
    `check_addr` : address offset check
    `check_field` : field check
    """
    def __init__(self, worksheets:dict[str,Worksheet]):
        self.worksheets = worksheets
        self.reg_model = {}

    def check_format(self):
        """
        检查基础格式, 包括表格项是否齐全, 位置是否正确, 特定单元格是否符合要求等.
        1. 基地址必须为十六进制并以0X(x)作前缀
        2. 地址偏移必须为十六进制并以0X(x)作前缀
        3. 支持的访问类型
        4. 寄存器位宽只能为32 bit或64 bit
        5. 域的比特位范围必须为xx:xx格式
        6. 复位值必须为十六进制并以0X(x)作前缀
        7. 域的同步复位信号可有多个

        此规则必须先于其他规则进行检查, 否则表征寄存器表结构的`self.reg_model`为空.

        Return
        ------
        `check_ack` : `bool`类型, Excel Workbook是否通过基础格式检查

        Note
        ----
        当格式检查通过, `self.reg_model`中就会包含解析出的表内容
        """
        self.reg_model = {}

        # 遍历每个Excel Workbook (Worksheet)进行检查
        for filename, worksheet in self.worksheets.items():
            base_row, last_row = 0, 1
            max_row = worksheet.max_row

            check_ack = True
            addrmap_name = os.path.split(filename)[-1].split(".")[0]
            self.reg_model[addrmap_name] = []

            while last_row < max_row:
                item_dict = {"BaseRow": base_row+1}
                last_rows = []

                # 验证每张表的表头: 表项名, 内容
                for key, val in EXCEL_REG_HEAD.items():
                    # 验证表项名是否正确
                    off_row, col = val["Entry"]["Loc"]
                    row = base_row + off_row
                    cell = worksheet.cell(row=row, column=col)

                    if cell.value not in val["Entry"]["Name"]:
                        message.error(
                            "file: %s, cell row: %d col: %d not match" % (filename, row, col),
                            raise_err = False
                        )
                        check_ack = False

                    # 验证表项内容是否合规
                    off_row, col = val["Content"]["Loc"]
                    row = base_row + off_row
                    cell = worksheet.cell(row=row, column=col)

                    if cell.value is None:
                        message.error(
                            "file: %s, cell row: %d col: %d cannot be empty" % (filename, row, col),
                            raise_err = False
                        )
                        check_ack = False
                    else:
                        # 对相应单元格进行正则匹配的检查
                        pattern = val["Content"].get("Pattern", None)
                        if pattern is not None:
                            # 验证正则表达式
                            if re.match(pattern, str(cell.value)) is None:
                                message.error(
                                    "file: %s, cell row: %d col: %d format not match" % (filename, row, col),
                                    raise_err = False
                                )
                                check_ack = False

                    if check_ack:
                        # 强制不转义字符串里的换行符, 避免之后生成RDL时格式错乱
                        # \r -> \\\\r, \n -> \\\\n
                        # Note: 实际上加了两级非转义, 因为RDL在编译的时候也会转义换行符
                        if val["Content"]["Format"] == "str":
                            item_dict[key] = cell.value.replace("\r", r"\\r").replace("\n", r"\\n")
                        else:
                            item_dict[key] = cell.value

                # 验证每张表中的域定义: 表项名, 内容
                for key, val in EXCEL_REG_FIELD.items():
                    item_dict[key] = []
                    # 验证域定义名是否正确: 比特位、域名称等
                    off_row, col = val["Entry"]["Loc"]
                    row = base_row + off_row
                    cell = worksheet.cell(row=row, column=col)

                    if cell.value not in val["Entry"]["Name"]:
                        message.error(
                            "file: %s, cell row: %d col: %d not match" % (filename, row, col),
                            raise_err = False
                        )
                        check_ack = False

                    # 验证各项域定义内容是否符合规则
                    off_row, col = val["Content"]["StartLoc"]
                    row = base_row + off_row
                    cell = worksheet.cell(row=row, column=col)

                    # 域定义不能为空，至少要写明所有bit为Reserved
                    if cell.value is None:
                        message.error(
                            "file: %s, cell row: %d col: %d cannot be empty, "
                            "if this register need to be empty, please mark all field bits as Reserved." % (filename, row, col),
                            raise_err = False
                        )
                        check_ack = False

                    while cell.value is not None:
                        # 对相应单元格进行正则匹配规则的检查
                        pattern = val["Content"].get("Pattern", None)
                        if pattern is not None:
                            # 验证正则表达式
                            if re.match(pattern, str(cell.value)) is None:
                                message.error(
                                    "file: %s, cell row: %d col: %d format not match" % (filename, row, col),
                                    raise_err = False
                                )
                                check_ack = False

                        if check_ack:
                            # 强制不转义字符串里的换行符, 避免之后生成RDL时格式错乱
                            # \r -> \\\r, \n -> \\\n
                            # Note: 实际上加了两级非转义, 因为RDL在编译的时候也会转义换行符
                            if val["Content"]["Format"] == "str":
                                item_dict[key].append(cell.value.replace("\r", r"\\r").replace("\n", r"\\n"))
                            else:
                                item_dict[key].append(cell.value)

                        row += 1
                        cell = worksheet.cell(row=row, column=col)

                    last_rows.append(row)

                # 保证域定义中每行都是完整的
                if not min(last_rows) == max(last_rows):
                    message.error(
                        "file: %s, base row: %d, all field items should not be empty" % (filename, base_row),
                        raise_err = False
                    )
                    check_ack = False
                else:
                    last_row = last_rows[0]

                # 每张表的行基数是开头的前一行
                # 表之间有空行间隔, 需要确定下一张表的基准位置
                empty_row = last_row
                while worksheet.cell(row=empty_row, column=1).value is None and empty_row <= max_row:
                    empty_row += 1
                base_row = empty_row - 1

                # 如果当前寄存器定义Table出现问题, 则停止检查后面的Table,
                # 直接返回False, 避免报错信息过多及影响后续格式检查
                if not check_ack:
                    return check_ack

                self.reg_model[addrmap_name].append(item_dict)

        self._format_regs(self.reg_model)

        return check_ack

    @property
    def parsed_model(self):
        """
        """
        assert self.reg_model
        return self.reg_model

    @staticmethod
    def _format_regs(reg_model:dict[str,list]):
        """
        Format the register model

        Parameter
        ---------
        `reg_model` :
        """
        for addrmap_entry in reg_model.values():
            for reg in addrmap_entry:
                for key, val in EXCEL_REG_HEAD.items():
                    if val["Content"]["Format"] == "hex":
                        reg[key] = int(reg[key], base=16)

                    elif val["Content"]["Format"] == "int":
                        if isinstance(reg[key], str):
                            reg[key] = int(reg[key], base=10)

                for key, val in EXCEL_REG_FIELD.items():
                    for idx, field_val in enumerate(reg[key]):
                        if val["Content"]["Format"] == "hex":
                            reg[key][idx] = int(field_val, base=16)

                        elif val["Content"]["Format"] == "int":
                            if isinstance(field_val, str):
                                reg[key][idx] = int(field_val, base=10)

                        elif val["Content"]["Format"] == "range":
                            reg[key][idx] = tuple(map(int, field_val.split(":")))

    def check_name(self):
        """
        """
        check_ack = True

        for addrmap_name, addrmap_entry in self.reg_model.items():
            names_full, names_abbr = [], []
            for reg in addrmap_entry:
                base_row = reg["BaseRow"]

                if reg["RegAbbr"] in names_abbr:
                    message.warning("addrmap: %s, base row: %d, duplicate register abbreviation: %s"
                                    % (addrmap_name, base_row, reg["RegAbbr"]))
                    check_ack = False
                else:
                    names_abbr.append(reg["RegAbbr"])

                if reg["RegName"] in names_full:
                    message.warning("addrmap: %s, base row: %d, duplicate register name: %s"
                                    % (addrmap_name, base_row, reg["RegName"]))
                    check_ack = False
                else:
                    names_full.append(reg["RegName"])

        return check_ack

    def check_addr(self):
        """
        检查寄存器的地址偏移(Address Offset)分配是否合法

        Rule
        -----
        1. 按寄存器字节长度的整数倍进行地址对齐
        2. 在同一Excel Workbook (addrmap)中是否有地址空间重叠
        """
        check_ack = True

        for addrmap_name, addrmap_entry in self.reg_model.items():
            addrs = []

            for reg in addrmap_entry:
                base_row = reg["BaseRow"]
                byte_size = reg["RegWidth"] // 8

                if not (reg["AddrOffset"] % byte_size == 0):
                    message.error(
                        "addrmap: %s, register: %s(base row: %d) address offset does not match alignment of length %d" % (
                                addrmap_name, reg["RegName"], base_row, reg["RegWidth"]
                        ),
                        raise_err = False
                    )

                    check_ack = False

                addr_low = reg["AddrOffset"]
                addr_high = addr_low + byte_size - 1
                addrs.append((addr_low, addr_high, reg["RegName"]))

            # 做一遍排序确定是否有地址重叠
            addrs.sort(key=lambda x:x[0])

            for idx in range(len(addrs) - 1):
                if addrs[idx][1] >= addrs[idx+1][0]:
                    message.warning(
                        "addrmap: %s, address space overlap between reg: %s(%s:%s) and reg: %s(%s:%s)" % (
                            addrmap_name, addrs[idx][2], hex(addrs[idx][0]), hex(addrs[idx][1]),
                            addrs[idx+1][2], hex(addrs[idx+1][0]), hex(addrs[idx+1][1])
                        )
                    )

                    check_ack = False

        return check_ack

    def check_field(self):
        """
        检查域定义

        Rule
        ----
        1. 比特位序从最高位起
        2. 域比特位范围按[high_bit]:[low_bit]排布
        3. 域比特位无重叠(3.1), 无遗漏(3.2)
        4. 复位值合法, 不能超过该域能表示的最大值
        5. 除Reserved以外域名称无重复
        6. Reserved域同步复位信号只能为None

        Return
        ------
        `bool`, whether field definitions are legal
        """
        for addrmap_name, addrmap_entry in self.reg_model.items():
            for reg in addrmap_entry:
                bit_size = reg["RegWidth"]
                reg_name = reg["RegName"]
                base_row = reg["BaseRow"]

                # Rule 2
                for bits in reg["FieldBit"]:
                    if bits[0] < bits[1]:
                        message.error(
                            "addrmap: %s, reg: %s(base row: %d), field %d:%d should be ordered from high to low, like %d:%d" % (
                                addrmap_name, reg_name, base_row, bits[0], bits[1], bits[1], bits[0]
                            ),
                            raise_err = False
                        )

                        return False

                for idx in range(len(reg["FieldBit"]) - 1):
                    fld, fld_next = reg["FieldBit"][idx], reg["FieldBit"][idx+1]

                    # Rule 3.1
                    if fld[1] <= fld_next[0]:
                        message.error(
                            "addrmap: %s, reg: %s(base row: %d), field %d:%d and field %d:%d may overlap or be in wrong order" % (
                                addrmap_name, reg_name, base_row, fld[0], fld[1], fld_next[0], fld_next[1]
                            ),
                            raise_err = False
                        )

                        return False

                    # Rule 1
                    elif not fld[1] == fld_next[0] + 1:
                        message.error(
                            "addrmap: %s, reg: %s(base row: %d), gap between field %d:%d and field %d:%d" % (
                                addrmap_name, reg_name, base_row, fld[0], fld[1], fld_next[0], fld_next[1]
                            ),
                            raise_err = False
                        )

                        return False

                # Rule 3.2
                if not (reg["FieldBit"][0][0] == bit_size - 1 and reg["FieldBit"][-1][1] == 0):
                    message.error(
                        "addrmap: %s, reg: %s(base row: %d), field bits do not match register length, "
                        "fill empty bits with 'Reserved', or discard some field bits" % (
                            addrmap_name, reg_name, base_row
                        ),
                        raise_err = False
                    )

                    return False

                # Rule 4
                for cnt, opval in enumerate(reg["FieldRstVal"]):
                    bit_high, bit_low = reg["FieldBit"][cnt][0], reg["FieldBit"][cnt][1]

                    if opval >= (2 << (bit_high - bit_low)):
                        message.error(
                            "addrmap: %s, reg: %s(base row: %d), field %d:%d reset value exceeds" % (
                                addrmap_name, reg_name, base_row, bit_high, bit_low
                            ),
                            raise_err = False
                        )

                        return False

                # Rule 5
                fld_names = set()

                for fld_name in reg["FieldName"]:
                    if fld_name not in fld_names or fld_name.lower() == "reserved":
                        fld_names.add(fld_name)

                    else:
                        message.error("addrmap: %s, reg: %s(base row: %d) has duplicate field name: %s" % (
                                addrmap_name, reg_name, base_row, fld_name
                            ),
                            raise_err = False
                        )

                        return False

        return True

class Parser:
    """
    """
    def __init__(self):
        self.rdlc = RDLCompiler()
        self.ipxact = IPXACTImporter(self.rdlc)

    def parse(self, orig_files:list, list_file:str, gen_dir:str,
            to_generate_rdl=False, excel_top="top_map"):
        """
        parse Excel worksheets, SystemRDL and IP-XACT files

        Parameter
        ---------
        `orig_files` : all input file names
        `list_file` : input file list name
        `gen_dir` : directory to save SystemRDL converted from Excel worksheets
        `to_generate_rdl` : whether to generate SystemRDL from Excel worksheets, dafaults to `False`
        `excel_top` :
            when `to_generate_rdl == True` and `orig_files` only consists of Excel Worksheet,
            this parameter specifies definition and instance name of top addrmap, defaults to `top_map`

        Calls
        -----
        `parse_excel`
        `parse_rdl_ipxact`
        """
        ori_rdl_ipxact_files, ori_excel_files = [], []

        all_files = []

        # use -f/--file option
        if orig_files is not None:
            if list_file is not None:
                message.warning("cannot use -f/--file and -l/--list options at the same time,"
                                "-l/--list option will be ignored")

            for file in orig_files:
                    if not os.path.exists(file):
                        message.error(
                            "input file: %s does not exists" % (file)
                        )

                    ff = os.path.splitext(file)[-1]
                    if ff in (".rdl", ".xml", ".xlsx"):
                        all_files.append(file)
                        if ff in (".rdl", ".xml"):
                            ori_rdl_ipxact_files.append(file)
                        elif ff == ".xlsx":
                            ori_excel_files.append(file)
                    else:
                        message.error(
                            "wrong file format (should be .rdl/.xml/.xlsx)"
                        )

        # use -l/--list option
        elif list_file is not None:
            if not os.path.exists(list_file):
                message.error(
                    "list file %s does not exists" % (
                        list_file
                    )
                )

            else:
                with open(list_file, "r") as f:
                    lines = f.readlines()

                    for cnt, line in enumerate(lines):

                        # comments(#) or blank lines
                        line = line.strip().replace("\n", "").replace("\r", "")

                        if line == "" or line.startswith("#"):
                            continue

                        if not os.path.exists(line):
                            message.error(
                                "list file: %s line: %d input file: %s does not exists" % (
                                    list_file, cnt, line
                                )
                            )

                        ff = os.path.splitext(line)[-1]

                        if ff in (".rdl", ".xml", ".xlsx"):
                            all_files.append(line)

                            if ff in (".rdl", ".xml"):
                                ori_rdl_ipxact_files.append(line)

                            elif ff == ".xlsx":
                                ori_excel_files.append(line)
                        else:
                            message.error(
                                "wrong file format (should be .rdl/.xml/.xlsx)"
                            )

        else:
            message.error(
                "one of the -f/--file and -l/--list options must be provided"
            )

        # 先解析Excel Worksheets生成中间RDL files,
        # 再合并其他输入的RDL代码一起引入systemrdl-compiler编译生成寄存器模型
        excel_rdl_files = self.parse_excel(
            files = ori_excel_files,
            top_name = excel_top,
            gen_rdl_dir = gen_dir,
            generate_rdl = to_generate_rdl,
            gen_extra_top = (ori_rdl_ipxact_files == [])
        )

        if excel_rdl_files is None:
            if to_generate_rdl:
                message.info(
                    "input files do not consist of Excel worksheet(.xlsx) files"
                )

            if not ori_rdl_ipxact_files:
                return None

        else:
            # 把原来的Excel files替换为生成后的RDL files
            f_iter = iter(excel_rdl_files)

            for idx, file in enumerate(all_files):
                if file.endswith(".xlsx"):
                    all_files[idx] = next(f_iter)

            # 如果输入全是Excel Worksheet,
            # 则生成的excel_rdl_files的最后会有一个包含top_addrmap的RDL file,
            # 需要加入到RDL parse的列表中去
            if not ori_rdl_ipxact_files:
                all_files.append(next(f_iter))

        return self.parse_rdl_ipxact(all_files)

    def parse_excel(self, files:list[str], top_name:str, gen_rdl_dir:str,
                    generate_rdl=False, gen_extra_top=False):
        """
        Parameter
        ---------
        `files` : 包含所有需要检查的Excel文件名的list
        `top_name` : `gen_extra_top == True`时生成的额外顶层addrmap及RDL文件名
        `gen_rdl_dir` : SystemRDL的生成目录
        `generate_rdl` : 是否生成SystemRDL

        Return
        ------
        `rdl_file` : `generate_rdl == True`时生成的多个或聚合的一个RDL文件名
            - `gen_aggregation == False`时为`list`类型
            - `gen_aggregation == True`时为`str`类型

        `None` : 输入`files`为空或者`generate_rdl == False`时无返回值
        """
        if not files:
            return None

        worksheets = {}

        for file in files:
            wb = load_workbook(file)
            ws = wb.active
            worksheets[file] = ws

        parser = ExcelParser(worksheets)

        for checker in [
            parser.check_format,
            parser.check_name,
            parser.check_addr,
            parser.check_field
        ]:
            if not checker():
                # raises error after checking all worksheets
                message.error(
                    "parser aborted due to previous errors"
                )

        message.info(
            "all Excel worksheets (.xlsx) are properly parsed"
        )

        if generate_rdl:
            from .excel.excel2rdl import RDLGenerator

            if not os.path.exists(gen_rdl_dir):
                message.error(
                    "invalid directory for the generated rdl file"
                )

            generator = RDLGenerator(parser.parsed_model)

            rdl_files = generator.generate_rdl(
                gen_rdl_dir,
                top_name,
                gen_extra_top
            )

            return rdl_files
        else:
            return None

    def parse_rdl_ipxact(self, files:list[str]):
        """
        Import and compile SystemRDL and IP-XACT files

        Parameter
        ---------
        `files` : all SystemRDL and IP-XACT files

        Return
        ------
        `root` : `systemrdl.node.RootNode`
        `None` : if `files` is empty
        """
        if files == []:
            return None

        # pre-defined SystemRDL files, such as user-defined properties
        predef_files = [
            os.path.join(os.path.dirname(__file__),  "user_def_property.rdl"),
            os.path.join(os.path.dirname(__file__),  "db_reg_def.rdl"),
            os.path.join(os.path.dirname(__file__),  "ras_reg_def.rdl")
        ]

        try:
            # compile pre-defined SystemRDL files
            for file in predef_files:
                self.rdlc.compile_file(file)

            # compile user-written input SystemRDL files
            for file in files:
                message.info(
                    "SystemRDL Compiler: import and compile %s" % (file)
                )

                if file.endswith(".xml"):
                    self.ipxact.import_file(file)
                else:
                    self.rdlc.compile_file(file)

            message.info("SystemRDL Compiler: start elaborating")

            root = self.rdlc.elaborate()
        except RDLCompileError:
            message.error(
                "SystemRDL Compiler: aborted due to previous errors"
            )
        else:
            message.info(
                "SystemRDL Compiler: all files are parsed successfully"
            )

        return root
