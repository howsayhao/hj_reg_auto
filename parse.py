import re
import sys
import os.path

import utils.message as message
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from systemrdl import RDLCompileError, RDLCompiler, RDLWalker

from excel.args import EXCEL_REG_FIELD, EXCEL_REG_HEAD


class ExcelParser:
    """
    解析并检查以Excel表格编写的寄存器Spec.

    `check_format` : 基础格式是否正确
    `check_name` : 寄存器是否重名
    `check_addr` : 基地址和偏移合法性
    `check_field` : 域定义重叠、遗漏、重名、复位值合法性等
    """
    def __init__(self, worksheets:dict[str,Worksheet]):
        self.worksheets = worksheets
        self.regs = {}

    def check_format(self):
        """
        检查基础格式,包括表格项是否齐全,位置是否正确,特定单元格是否符合要求等.
        0. 每张寄存器规格表之间间隔为两行
        1. 基地址必须为十六进制并以0X(x)作前缀
        2. 地址偏移必须为十六进制并以0X(x)作前缀
        3. 支持的访问类型
        4. 寄存器长度只能为32或64bit
        5. 域的比特位范围必须为xx:xx格式
        6. 可选值的格式必须为n1[,n2,...], n1~n2, Others, All
        7. 复位值必须为十六进制并以0X(x)作前缀

        此规则必须先于其他规则进行检查, 否则表征寄存器表结构的`self.regs`为空.

        Return
        ------
        bool : 格式检查正确与否
               当格式检查通过, `self.regs`中就会包含解析出的表内容
        """
        # 防止多次调用该方法时重复写入相同内容
        self.regs = {}
        # 逐张worksheet进行检查
        for filename, worksheet in self.worksheets.items():
            base_row, last_row = 0, 1
            max_row = worksheet.max_row

            check_ack = True
            rfile_name = os.path.split(filename)[-1].split(".")[0]
            self.regs[rfile_name] = []

            while last_row < max_row:
                item_dict = {"BaseRow": base_row+1}
                last_rows = []

                for key, val in EXCEL_REG_HEAD.items():
                    # 验证表项名是否正确
                    off_row, col = val["Entry"]["Loc"]
                    row = base_row + off_row
                    cell = worksheet.cell(row=row, column=col)

                    if cell.value not in val["Entry"]["Name"]:
                        message.error("file:%s, cell row:%d col:%d not match!" %(filename, row, col))
                        check_ack = False

                    # 验证表项内容是否符合规则
                    off_row, col = val["Content"]["Loc"]
                    row = base_row + off_row
                    cell = worksheet.cell(row=row, column=col)

                    if cell.value is None:
                        message.error("file:%s, cell row:%d col: %d cannot be empty!" %(filename, row, col))
                        check_ack = False
                    else:
                        # 对相应单元格进行正则匹配规则的检查
                        pattern = val["Content"]["Pattern"]
                        if pattern is not None:
                            # 验证正则表达式
                            if re.match(pattern, str(cell.value)) is None:
                                message.error("file:%s, cell row:%d col:%d format not match!" %(filename, row, col))
                                check_ack = False

                    if check_ack:
                        # 强制不转义字符串里的换行符, 避免之后生成RDL时格式错乱
                        # \r -> \\\r, \n -> \\\n
                        # Note: 实际上加了两级非转义, 因为RDL在编译的时候也会转义换行符
                        if val["Content"]["Format"] == "str":
                            item_dict[key] = cell.value.replace("\r", r"\\r").replace("\n", r"\\n")
                        else:
                            item_dict[key] = cell.value

                # 验证表的域定义项: 包括表项名和内容
                for key, val in EXCEL_REG_FIELD.items():
                    item_dict[key] = []
                    # 验证域定义名是否正确: 比特位、域名称等
                    off_row, col = val["Entry"]["Loc"]
                    row = base_row + off_row
                    cell = worksheet.cell(row=row, column=col)

                    if cell.value not in val["Entry"]["Name"]:
                        message.error("file:%s, cell row:%d col:%d not match!" %(filename, row, col))
                        check_ack = False

                    # 验证各项域定义内容是否符合规则
                    off_row, col = val["Content"]["StartLoc"]
                    row = base_row + off_row
                    cell = worksheet.cell(row=row, column=col)

                    # 域定义不能为空，至少要写明所有bit为Reserved
                    if cell.value is None:
                        message.error("file:%s, cell row:%d col: %d cannot be empty!" %(filename, row, col))
                        check_ack = False
                        message.info("if this register need to be empty, please mark all field bits as Reserved.")

                    while cell.value is not None:
                        # 对相应单元格进行正则匹配规则的检查
                        pattern = val["Content"]["Pattern"]
                        if pattern is not None:
                            # 验证正则表达式
                            if re.match(pattern, str(cell.value)) is None:
                                message.error("file:%s, cell row:%d col:%d format not match!" %(filename, row, col))
                                check_ack = False

                        if check_ack:
                            # 强制不转义字符串里的换行符, 避免之后生成RDL时格式错乱
                            # \r -> \\\r, \n -> \\\n
                            # Note: 实际上加了两级非转义, 因为RDL在编译的时候也会转义换行符
                            if val["Content"]["Format"] == "str":
                                item_dict[key].append(cell.value.replace("\r", r"\\r").replace("\n", r"\\n"))
                            else:
                                item_dict[key].append(cell.value)

                        row += 1
                        cell = worksheet.cell(row=row, column=col)

                    last_rows.append(row)

                # 保证域定义中每行都是完整的
                if not min(last_rows) == max(last_rows):
                    message.error("file:%s, base row:%d, all field items should not be empty!"
                                  % (filename, base_row))
                    check_ack = False
                else:
                    last_row = last_rows[0]

                # 每张表的行基数是开头的前一行
                # 表之间有空行间隔, 需要确定下一张表的基准位置
                empty_row = last_row
                while worksheet.cell(row=empty_row, column=1).value is None and empty_row <= max_row:
                    empty_row += 1
                base_row = empty_row - 1

                # 如果当前寄存器定义Table出现问题, 则停止检查后面的Table,
                # 直接返回False, 避免报错信息过多及影响后续格式检查
                if not check_ack:
                    return check_ack

                self.regs[rfile_name].append(item_dict)

        self.format_regs(self.regs)

        return check_ack

    @property
    def parsed_model(self):
        """
        解析完成的寄存器架构
        """
        assert self.regs != {}
        return self.regs

    @staticmethod
    def format_regs(regs:dict[str,list]):
        for rfile_entry in regs.values():
            for reg in rfile_entry:
                for key, val in EXCEL_REG_HEAD.items():
                    if val["Content"]["Format"] == "hex":
                        reg[key] = int(reg[key], base=16)
                    elif val["Content"]["Format"] == "int":
                        if isinstance(reg[key], str):
                            reg[key] = int(reg[key], base=10)

                for key, val in EXCEL_REG_FIELD.items():
                    for idx, field_val in enumerate(reg[key]):
                        if val["Content"]["Format"] == "hex":
                            reg[key][idx] = int(field_val, base=16)
                        elif val["Content"]["Format"] == "int":
                            if isinstance(field_val, str):
                                reg[key][idx] = int(field_val, base=10)
                        elif val["Content"]["Format"] == "range":
                            reg[key][idx] = tuple(map(int, field_val.split(":")))

    def check_name(self):
        """
        检查寄存器名称和简写是否有重复

        Note
        ----
        该规则跨文件进行检查, 请保证所有文件中的寄存器名和简写均无重复

        TO BE DONE
        ----------
        提供别名寄存器(Alias)的支持, 为相同寄存器提供多个物理地址和定义
        """
        names_full, names_abbr = [], []
        check_ack = True
        for rfile_name, rfile_entry in self.regs.items():
            for reg in rfile_entry:
                base_row = reg["BaseRow"]

                if reg["RegAbbr"] in names_abbr:
                    message.warning("regfile:%s, base row:%d, duplicate register abbreviation:%s"
                                    % (rfile_name, base_row, reg["RegAbbr"]))
                    check_ack = False
                else:
                    names_abbr.append(reg["RegAbbr"])

                if reg["RegName"] in names_full:
                    message.warning("regfile:%s, base row:%d, duplicate register name:%s"
                                    % (rfile_name, base_row, reg["RegName"]))
                    check_ack = False
                else:
                    names_full.append(reg["RegName"])

        return check_ack

    def check_addr(self):
        """
        检查寄存器的基地址和偏移是否合规.

        Rule
        -----
        1. 按寄存器字节长度的整数倍标定地址偏移 (Alignment)
        2. 同一文件(regfile)中是否有地址空间重叠 (Address space overlap)

        Return
        ------
        `bool` : `check_ack` 规则检查是否通过
        """
        check_ack = True

        for rfile_name, rfile_entry in self.regs.items():
            addrs = []

            for reg in rfile_entry:
                base_row = reg["BaseRow"]
                byte_size = reg["RegWidth"] // 8

                if not (reg["AddrOffset"] % byte_size == 0):
                    message.error("regfile:%s, register:%s(base row:%d) address offset does not match alignment of length %d"
                                % (rfile_name, reg["RegName"], base_row, reg["RegWidth"]))
                    check_ack = False
                addr_low = reg["AddrOffset"]
                addr_high = addr_low + byte_size - 1
                addrs.append((addr_low, addr_high, reg["RegName"]))

            # 做一遍排序确定是否有地址重叠
            addrs.sort(key=lambda x:x[0])
            for idx in range(len(addrs)-1):
                if addrs[idx][1] >= addrs[idx+1][0]:
                    message.warning("regfile:%s, address space overlap between reg:%s(%s:%s) and reg:%s(%s:%s)"
                                    % (rfile_name, addrs[idx][2], hex(addrs[idx][0]), hex(addrs[idx][1]),
                                       addrs[idx+1][2], hex(addrs[idx+1][0]), hex(addrs[idx+1][1])))
                    check_ack = False

        return check_ack

    def check_field(self):
        """
        检查域定义
        
        Rule
        ----
        1. 比特位序从最高位起
        2. 域比特位范围按[high_bit]:[low_bit]排布
        3. 域比特位无重叠(3.1), 无遗漏(3.2)
        4. 复位值合法, 不能超过该域能表示的最大值
        5. 除Reserved以外域名称无重复
        """
        for rfile_name, rfile_entry in self.regs.items():
            for reg in rfile_entry:
                bit_size = reg["RegWidth"]
                reg_name = reg["RegName"]
                base_row = reg["BaseRow"]

                # Rule 2
                for bits in reg["FieldBit"]:
                    if bits[0] < bits[1]:
                        message.error("regfile:%s, reg:%s(base row:%d), field %d:%d should be ordered from high to low, like %d:%d"
                                      % (rfile_name, reg_name, base_row, bits[0], bits[1], bits[1], bits[0]))
                        return False

                for idx in range(len(reg["FieldBit"])-1):
                    fld, fld_next = reg["FieldBit"][idx], reg["FieldBit"][idx+1]
                    # Rule 3.1
                    if fld[1] <= fld_next[0]:
                        message.error("regfile:%s, reg:%s(base row:%d), field %d:%d and field %d:%d may overlap or be in wrong order"
                                    % (rfile_name, reg_name, base_row, fld[0], fld[1], fld_next[0], fld_next[1]))
                        return False
                    # Rule 1
                    elif not fld[1] == fld_next[0] + 1:
                        message.error("regfile:%s, reg:%s(base row:%d), gap between field %d:%d and field %d:%d"
                                      % (rfile_name, reg_name, base_row, fld[0], fld[1], fld_next[0], fld_next[1]))
                        return False

                # Rule 3.2
                if not (reg["FieldBit"][0][0] == bit_size - 1 and reg["FieldBit"][-1][1] == 0):
                    message.error("regfile:%s, reg:%s(base row:%d), field bits do not match register length, "
                                  "fill empty bits with 'Reserved', or discard some field bits"
                                  % (rfile_name, reg_name, base_row))
                    return False

                # Rule 4
                for cnt, opval in enumerate(reg["FieldRstVal"]):
                    bit_high, bit_low = reg["FieldBit"][cnt][0], reg["FieldBit"][cnt][1]
                    if opval >= (2 << (bit_high - bit_low)):
                        message.error("regfile:%s, reg:%s(base row:%d), field %d:%d reset value exceeds"
                                      % (rfile_name, reg_name, base_row, bit_high, bit_low))
                        return False

                # Rule 5
                fld_names = set()
                for fld_name in reg["FieldName"]:
                    if fld_name not in fld_names or fld_name.lower() == "reserved":
                        fld_names.add(fld_name)
                    else:
                        message.error("regfile:%s, reg:%s(base row:%d) has duplicate field name:%s" 
                                      % (rfile_name, reg_name, base_row, fld_name))
                        return False

        return True


def parse_excel(files:list[str], generate_rdl:bool, module_name:str, rdl_dir:str):
    """
    Parameter
    ---------
    `files` : 包含所有需要检查的Excel文件名的list
    `generate_rdl` : 是否生成SystemRDL
    `module_name` : 生成的SystemRDL的模块名(addrmap名)
    `rdl_dir` : SystemRDL的生成路径

    Return
    ------
    No return

    """
    worksheets = {}
    for file in files:
        wb = load_workbook(file)
        ws = wb.active
        worksheets[file] = ws
    
    parser = ExcelParser(worksheets)

    for checker in [parser.check_format,
                    parser.check_name,
                    parser.check_addr,
                    parser.check_field]:
        if not checker():
            message.info("parser aborted due to previous error")
            sys.exit(1)

    message.info("all files have been parsed successfully")
    if generate_rdl:
        from .excel.gen_rdl import RDLGenerator

        if not os.path.exists(rdl_dir):
            message.error("specified an invalid path for the generated rdl file!")
            sys.exit(1)

        generator = RDLGenerator(reg_model=parser.parsed_model,
                                 gen_dir=rdl_dir,
                                 module_name=module_name)
        generator.generate_rdl()
        parse_rdl([os.path.join(rdl_dir, module_name + ".rdl")])

def show_excel_rules():
    """
    显示Excel解析需要遵循的规则
    """
    print(ExcelParser.__doc__, "\n",
          ExcelParser.check_format.__doc__, "\n",
          ExcelParser.check_name.__doc__, "\n",
          ExcelParser.check_addr.__doc__, "\n",
          ExcelParser.check_field.__doc__)

def parse_rdl(files:list[str]):
    """
    使用`systemrdl-compiler`编译并解析SystemRDL文件

    Parameter
    ---------
    `files` : 需要解析的文件名, 以`list`形式组织
    """
    rdlc = RDLCompiler()

    try:
        for file in files:
            rdlc.compile_file(file)

        root = rdlc.elaborate()
    except RDLCompileError:
        sys.exit(1)
    else:
        message.info("SystemRDL parsed successfully")
