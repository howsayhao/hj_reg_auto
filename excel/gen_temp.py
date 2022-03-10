import os.path
from shutil import copy
from .args import EXCEL_REG_HEAD, EXCEL_REG_FIELD

import utils.message as message
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange


def generate_excel(path:str, name:str, rnum:int, rname:list, language:str, table_interval=4):
    """
    以单个寄存器模板为基础生成的寄存器组(regfile)的规格说明

    Parameter
    ---------
    `path` : 生成模板存放的路径
    `name` : 生成模板的名字, 后续将作为regfile的名称使用
    `rnum` : 生成模板中寄存器的数量
    `rname` : 生成模板中各个寄存器的名字
    `language` : 生成模板的语言格式, 支持中文/English

    Return
    ------
    No return
    """
    temp_file = os.path.join("excel", "templates", "template_{}.xlsx".format(language))
    gen_file = os.path.join(path, name)

    # 重名处理
    if os.path.exists(gen_file):
        prefix_num = 1
        gen_name = str(prefix_num) + "_" + name
        while os.path.exists(os.path.join(path, gen_name)):
            prefix_num += 1
            gen_name = str(prefix_num) + "_" + name
        gen_file = os.path.join(path, gen_name)
        message.info("same template file name already exists, generate %s" % (gen_file))

    copy(temp_file, gen_file)

    if rnum > 1:
        wb = load_workbook(gen_file)
        ws = wb.active
        # 代表每张寄存器说明表单的长宽规格
        row_num, col_num = ws.max_row, ws.max_column
        source_cells = list(ws.rows)

        for row in range(1, row_num + 1):
            for col in range(1, col_num + 1):
                source_cell = source_cells[row-1][col-1]

                for regidx in range(rnum):
                    new_row = row + (row_num + table_interval) * regidx
                    new_cell = ws.cell(row=new_row, column=col)

                    if regidx >= 1:
                        # 复制值和格式
                        new_cell.value = source_cell.value
                        new_cell._style = source_cell._style
                        new_cell.number_format = source_cell.number_format

                    # 更新每张表的寄存器名和地址偏移
                    if (row, col) == EXCEL_REG_HEAD["RegName"]["Content"]["Loc"]:
                        new_cell.value = rname[regidx].upper()
                    elif (row, col) == EXCEL_REG_HEAD["RegAbbr"]["Content"]["Loc"]:
                        new_cell.value = rname[regidx].upper()
                    elif (row, col) == EXCEL_REG_HEAD["AddrOffset"]["Content"]["Loc"]:
                        new_cell.value = "{0:#0{1}X}".format(regidx<<2, 10)

        # FIX: 上面Style Copy存在不能复制被合并的单元格的问题,
        # 原因是在openpyxl里单元格合并不是一个Style,
        # 而是将被合并单元格变成MergedCell Object,
        # 与普通的Cell Object不同
        area = CellRange(min_row=1, max_row=row_num, min_col=1, max_col=col_num)
        for mcr in ws.merged_cells:
            if mcr.coord not in area:
                continue
            cr = CellRange(mcr.coord)
            for regidx in range(1, rnum):
                cr.shift(row_shift=row_num+table_interval) 
                ws.merge_cells(cr.coord)

        wb.save(gen_file)

def genrate_rdl():
    pass
