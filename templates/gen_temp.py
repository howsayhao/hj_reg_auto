import os.path
from shutil import copy

import jinja2 as jj
import utils.message as message
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange
from parsers.excel.args import EXCEL_REG_HEAD


def generate_excel_template(dir:str, name:str, rnum:int, rname:list, language:str, table_interval=4):
    """
    generate templates in Excel worksheet (.xlsx) format

    Parameter
    ---------
    `dir` : directory to generated template file
    `name` : name of generated template file
    `rnum` : register number
    `rname` : register name list
    `language` : cn or en
    """
    temp_file = os.path.join(os.path.dirname(__file__), "template_{}.xlsx".format(language))
    gen_file = os.path.join(dir, name)

    # handle duplicate file names
    suffix_num = 1
    has_duplicate = False
    while os.path.exists(os.path.join(dir, name)):
        has_duplicate = True
        name = "{}_{}{}".format(os.path.splitext(name)[0], suffix_num, os.path.splitext(name)[1])
        suffix_num += 1

    gen_file = os.path.join(dir, name)
    copy(temp_file, gen_file)

    if rnum > 1:
        wb = load_workbook(gen_file)
        ws = wb.active
        # length and width of each register table
        row_num, col_num = ws.max_row, ws.max_column
        source_cells = list(ws.rows)

        for row in range(1, row_num + 1):
            for col in range(1, col_num + 1):
                source_cell = source_cells[row-1][col-1]

                for regidx in range(rnum):
                    new_row = row + (row_num + table_interval) * regidx
                    new_cell = ws.cell(row=new_row, column=col)

                    if regidx >= 1:
                        # copy values and formats
                        new_cell.value = source_cell.value
                        new_cell._style = source_cell._style
                        new_cell.number_format = source_cell.number_format

                    # update register name and address offset of each table
                    if (row, col) == EXCEL_REG_HEAD["RegName"]["Content"]["Loc"]:
                        new_cell.value = rname[regidx].upper()
                    elif (row, col) == EXCEL_REG_HEAD["RegAbbr"]["Content"]["Loc"]:
                        new_cell.value = rname[regidx].upper()
                    elif (row, col) == EXCEL_REG_HEAD["AddrOffset"]["Content"]["Loc"]:
                        new_cell.value = "{0:#0{1}X}".format(regidx<<2, 10)

        # fix: merged cells
        area = CellRange(min_row=1, max_row=row_num, min_col=1, max_col=col_num)
        for mcr in ws.merged_cells:
            if mcr.coord not in area:
                continue
            cr = CellRange(mcr.coord)
            for regidx in range(1, rnum):
                cr.shift(row_shift=row_num+table_interval)
                ws.merge_cells(cr.coord)

        wb.save(gen_file)

    if has_duplicate:
        message.warning("same template file name already exists, generate %s" % (gen_file))
    else:
        message.info("generate template: %s" % (gen_file))

def generate_rdl_template(dir:str, name:str, template_type:str, **kwargs):
    """
    generate templates in SystemRDL (.rdl) format

    template_type: common, interrupt
    """
    common_template_file = os.path.join(os.path.dirname(__file__), "template.rdl")

    # handle duplicate file names
    suffix_num = 1
    original_name = name

    while os.path.exists(os.path.join(dir, name)):
        name = "{}_{}{}".format(
            os.path.splitext(original_name)[0],
            suffix_num,
            os.path.splitext(original_name)[1]
        )
        suffix_num += 1

    gen_file = os.path.join(dir, name)

    # use jinja2 engine to dump common interrupt and ras arch templates
    jj_env = jj.Environment(
        loader=jj.FileSystemLoader(os.path.dirname(__file__)))

    if template_type == "common":
        copy(common_template_file, gen_file)

    elif template_type == "interrupt":
        template = jj_env.get_template("intr.jinja")

        intr_num = kwargs.pop("intr_num")
        stream = template.stream(intr_num=intr_num, inst_name=os.path.splitext(name)[0])
        stream.dump(os.path.join(dir, name))

    elif template_type == "ras":
        template = jj_env.get_template("ras.jinja")

        ras_record_list = kwargs.pop("ras_record_list")
        stream = template.stream(record_list=ras_record_list, inst_name=os.path.splitext(name)[0])
        stream.dump(os.path.join(dir, name))

    if kwargs:
        message.info("some arguments are not valid when generating templates")

    message.info("generate template: %s" % (gen_file))